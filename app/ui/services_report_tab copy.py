from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QTableWidget, QTableWidgetItem, QPushButton, QLabel, QHeaderView, QDialog, QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt
from app.core.db import get_all_guests, get_guest_services, get_all_checkins
from app.ui.dialogs.add_extra_service_dialog import AddExtraServiceDialog
from app.ui.dialogs.view_guest_services_dialog import ViewGuestServicesDialog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font, NamedStyle
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.dimensions import SheetFormatProperties
from decimal import Decimal

class ServicesReportTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Global Invoice and Export Buttons
        btn_layout = QHBoxLayout()
        global_invoice_btn = QPushButton("Generate Global Invoice")
        global_invoice_btn.setObjectName("actionButton")
        global_invoice_btn.clicked.connect(self.generate_global_invoice)
        btn_layout.addWidget(global_invoice_btn)
        export_xlsx_btn = QPushButton("Export as XLSX")
        export_xlsx_btn.setObjectName("actionButton")
        export_xlsx_btn.clicked.connect(self.export_as_xlsx)
        btn_layout.addWidget(export_xlsx_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # Search bar
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search guests by name...")
        self.search_input.textChanged.connect(self.search_guests)
        search_layout.addWidget(QLabel("Search:"))
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)

        # Guests table
        self.guest_table = QTableWidget()
        self.guest_table.setColumnCount(7)
        self.guest_table.setHorizontalHeaderLabels([
            "Name", "ID Number", "Nationality", "Phone", "Email", "VIP Status", "Actions"
        ])
        self.guest_table.setAlternatingRowColors(True)
        self.guest_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.guest_table.verticalHeader().setDefaultSectionSize(50)
        layout.addWidget(self.guest_table)

        self.load_guests()

    def load_guests(self):
        guests = get_all_guests()
        self.all_guests = guests
        self.display_guests(guests)

    def display_guests(self, guests):
        self.guest_table.setRowCount(len(guests))
        for row, guest in enumerate(guests):
            name = f"{guest['first_name']} {guest['last_name']}"
            self.guest_table.setItem(row, 0, QTableWidgetItem(name))
            self.guest_table.setItem(row, 1, QTableWidgetItem(guest.get('id_number') or ""))
            self.guest_table.setItem(row, 2, QTableWidgetItem(guest.get('nationality') or ""))
            phone = f"{guest.get('phone_code') or ''} {guest.get('phone_number') or ''}".strip()
            self.guest_table.setItem(row, 3, QTableWidgetItem(phone))
            self.guest_table.setItem(row, 4, QTableWidgetItem(guest.get('email') or ""))
            self.guest_table.setItem(row, 5, QTableWidgetItem(guest.get('vip_status') or ""))

            # Actions
            actions_widget = QWidget()
            actions_layout = QHBoxLayout(actions_widget)
            actions_layout.setContentsMargins(0, 0, 0, 0)
            actions_layout.setSpacing(5)

            add_extras_btn = QPushButton("Add Extras")
            add_extras_btn.setObjectName("actionButton")
            add_extras_btn.clicked.connect(lambda _, g=guest: self.open_add_extras_dialog(g))
            view_btn = QPushButton("View")
            view_btn.setObjectName("actionButton")
            view_btn.clicked.connect(lambda _, g=guest: self.open_view_services_dialog(g))
            actions_layout.addWidget(add_extras_btn)
            actions_layout.addWidget(view_btn)
            actions_layout.addStretch()
            self.guest_table.setCellWidget(row, 6, actions_widget)

    def search_guests(self):
        search_text = self.search_input.text().lower()
        filtered = [g for g in self.all_guests if search_text in f"{g['first_name']} {g['last_name']}`".lower()]
        self.display_guests(filtered)

    def open_add_extras_dialog(self, guest):
        dialog = AddExtraServiceDialog(guest, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_guests()

    def open_view_services_dialog(self, guest):
        dialog = ViewGuestServicesDialog(guest, self)
        dialog.exec()

    def generate_global_invoice(self):
        # To be implemented: generate a PDF invoice for all guests' extra services
        pass

    def export_as_xlsx(self):
        # Gather all guest services
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="thick", color="000000")
        guests = get_all_guests()
        all_data = []
        checkins = get_all_checkins()
        for guest in guests:
            # Find latest check-in for this guest (if any)
            guest_checkins = [c for c in checkins if c.get('guest_id') == guest['id']]
            room_number = ''
            if guest_checkins:
                # Use the most recent check-in
                latest_checkin = max(guest_checkins, key=lambda c: c.get('checkin_date', ''))
                room_number = latest_checkin.get('room_number', '')
            services = get_guest_services(guest['id'])
            for s in services:
                all_data.append({
                    'guest_id': guest['id'],
                    'guest_name': f"{guest['first_name']} {guest['last_name']}",
                    'room_number': room_number,
                    'service_name': s.get('service_name', ''),
                    'quantity': s.get('quantity', 0),
                    'unit_price': s.get('unit_price_at_time_of_charge', 0),
                    'total': s.get('total_charge', 0),
                    'date': s.get('charge_date', ''),
                    'status': 'Paid' if s.get('is_paid', 0) else ('Partly Paid' if s.get('amount_paid', 0) else 'Unpaid'),
                    'notes': s.get('notes', ''),
                    'amount_paid': s.get('amount_paid', 0),
                    'remaining_amount': s.get('remaining_amount', 0)
                })
        if not all_data:
            QMessageBox.information(self, "Export", "No data to export.")
            return
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Guest Services"
        # Set margins
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0, footer=0)
        # Fit to 1 page wide, blank for tall
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 # This means fit to 1 page wide, and scale height as needed across multiple pages
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE # Changed to LANDSCAPE
        ws.page_setup.horizontalCentered = True
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10  # Room Number
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15  # Total Due
        ws.column_dimensions['I'].width = 15  # Remaining
        ws.column_dimensions['J'].width = 20  # Payment Status
        ws.column_dimensions['K'].width = 20  # Notes
        # Header row
        header = ["Guest Name", "Service Name", "Room Number", "Quantity", "Unit Price", "Total", "Date", "Total Due", "Remaining", "Payment Status", "Notes"]
        ws.append(header)
        # Style header
        fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        font = Font(bold=True)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Create named styles for payment status
        good_style = NamedStyle(name="good")
        good_style.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        good_style.font = Font(color="006100")  # Dark green text

        bad_style = NamedStyle(name="bad")
        bad_style.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        bad_style.font = Font(color="9C0006")  # Dark red text

        neutral_style = NamedStyle(name="neutral")
        neutral_style.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
        neutral_style.font = Font(color="9C5700")  # Dark orange text

        # Write data with merged guest name cells
        row_idx = 2  # Excel rows are 1-indexed, and row 1 is header
        i = 0
        date_style = NamedStyle(name="short_date")
        date_style.number_format = 'yyyy-mm-dd'
        thick_border_rows = []  # Track rows that need thick bottom borders

        while i < len(all_data):
            guest_id = all_data[i]['guest_id']
            # Count how many rows for this guest
            count = 1
            while i + count < len(all_data) and all_data[i + count]['guest_id'] == guest_id:
                count += 1

            # Calculate totals for this guest
            guest_services = [all_data[i + j] for j in range(count)]
            guest_total_due = sum(Decimal(str(s['total'])) for s in guest_services)

            # A payment transaction has occurred if any service for the guest has a non-zero amount_paid.
            # The DB logic stamps the same payment info (amount_paid, remaining_amount) on all services for a guest during a transaction.
            payment_has_occurred = any(Decimal(str(s.get('amount_paid', 0))) > 0 for s in guest_services)

            if payment_has_occurred:
                # If a payment was made, the 'remaining_amount' on the service records reflects the guest's total remaining balance.
                # All service records for that guest updated in the same transaction will have the same remaining_amount.
                # We can just take it from the first service record.
                guest_remaining = Decimal(str(guest_services[0].get('remaining_amount', '0')))
                guest_total_paid = guest_total_due - guest_remaining
            else:
                # No payment has ever been made for this guest.
                guest_total_paid = Decimal('0')
                guest_remaining = guest_total_due

            # Ensure remaining is not negative.
            guest_remaining = max(Decimal('0'), guest_remaining)
            guest_total_paid = max(Decimal('0'), guest_total_paid)

            # Determine overall payment status for this guest. Using a small epsilon for safety.
            if guest_remaining <= Decimal('0.01'):
                guest_payment_status = 'Paid'
            elif guest_total_paid > 0:
                guest_payment_status = 'Partly Paid'
            else:
                guest_payment_status = 'Unpaid'

            # Write rows
            for j in range(count):
                row = all_data[i + j]
                ws.append([
                    row['guest_name'] if j == 0 else '',
                    row['service_name'], row['room_number'], row['quantity'], row['unit_price'],
                    row['total'], row['date'],
                    f"{guest_total_due:.2f}" if j == 0 else '',  # Total Due (only in first row)
                    f"{guest_remaining:.2f}" if j == 0 else '',  # Remaining (only in first row)
                    guest_payment_status if j == 0 else '',  # Payment Status (only in first row)
                    row['notes']
                ])
                # Align guest name (A) left, service name (B) center
                ws.cell(row=row_idx + j, column=1).alignment = Alignment(horizontal="left", vertical="top")
                ws.cell(row=row_idx + j, column=2).alignment = Alignment(horizontal="center", vertical="center")

                # Apply payment status cell style to the merged cell
                if j == 0:  # Only for the first row of each guest
                    status_cell = ws.cell(row=row_idx + j, column=10)  # Payment Status column
                    if guest_payment_status == 'Paid':
                        status_cell.style = good_style
                    elif guest_payment_status == 'Unpaid':
                        status_cell.style = bad_style
                    elif guest_payment_status == 'Partly Paid':
                        status_cell.style = neutral_style

                # Format date cell as short date if possible
                date_cell = ws.cell(row=row_idx + j, column=7)
                try:
                    # Try to parse and set as date
                    import datetime
                    if row['date']:
                        date_obj = datetime.datetime.strptime(row['date'][:10], '%Y-%m-%d')
                        date_cell.value = date_obj
                        date_cell.style = date_style
                except Exception:
                    pass

                # Align Total Due and Remaining columns (center top)
                if j == 0:  # Only for the first row of each guest
                    ws.cell(row=row_idx + j, column=8).alignment = Alignment(horizontal="center", vertical="top")  # Total Due
                    ws.cell(row=row_idx + j, column=9).alignment = Alignment(horizontal="center", vertical="top")  # Remaining
                    ws.cell(row=row_idx + j, column=10).alignment = Alignment(horizontal="center", vertical="top")  # Payment Status

            # Merge guest name cells if more than one row
            if count > 1:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + count - 1, end_column=1)
                # Merge Total Due, Remaining, and Payment Status cells
                ws.merge_cells(start_row=row_idx, start_column=8, end_row=row_idx + count - 1, end_column=8)  # Total Due
                ws.merge_cells(start_row=row_idx, start_column=9, end_row=row_idx + count - 1, end_column=9)  # Remaining
                ws.merge_cells(start_row=row_idx, start_column=10, end_row=row_idx + count - 1, end_column=10)  # Payment Status

            # Mark the last row of this guest's group for thick border
            thick_border_rows.append(row_idx + count - 1)

            row_idx += count
            i += count

        # Apply borders - general borders first
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                # Thick border for outside, thin for inside
                border = Border(
                    left=thick if c == 1 else thin,
                    right=thick if c == ws.max_column else thin,
                    top=thick if r == 1 else thin,
                    bottom=thick if r == ws.max_row else thin
                )
                cell.border = border

        # Now apply thick borders under each guest's group
        for row_num in thick_border_rows:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col)
                border = cell.border
                cell.border = Border(
                    left=border.left,
                    right=border.right,
                    top=border.top,
                    bottom=thick
                )

        # Freeze top row
        ws.freeze_panes = ws["A2"]
        # Save file dialog
        file_path, _ = QFileDialog.getSaveFileName(self, "Save XLSX Report", "guest_services_report.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            wb.save(file_path)
            QMessageBox.information(self, "Export", f"Report exported to {file_path}")