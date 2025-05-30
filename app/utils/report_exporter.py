from fpdf import FPDF
from openpyxl import Workbook

def export_checkins_pdf(table_widget, file_path):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    col_count = table_widget.columnCount()
    headers = [table_widget.horizontalHeaderItem(i).text() for i in range(col_count)]
    col_widths = [190 / col_count] * col_count

    pdf.set_fill_color(200, 220, 255)
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, border=1, ln=0, align="C", fill=True)
    pdf.ln()

    row_count = table_widget.rowCount()
    for row in range(row_count):
        for col in range(col_count):
            item = table_widget.item(row, col)
            text = item.text() if item else ""
            pdf.cell(col_widths[col], 10, text, border=1, ln=0, align="C")
        pdf.ln()

    pdf.output(file_path)

def export_checkins_xlsx(table_widget, file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Check-ins Report"

    headers = [table_widget.horizontalHeaderItem(i).text() for i in range(table_widget.columnCount())]
    ws.append(headers)

    for row in range(table_widget.rowCount()):
        row_data = []
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            row_data.append(item.text() if item else "")
        ws.append(row_data)

    wb.save(file_path)
